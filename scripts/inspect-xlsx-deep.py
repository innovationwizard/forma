#!/usr/bin/env python3
"""Deep xlsx inspection — surfaces every Dirty-George gotcha in a workbook.

Per `reference_dirty_george_principle` memory: the structural manifest
(sheet names, dimensions, header rows, totals) is necessary but never
sufficient. Before writing a parser, we need to know about merged cells,
hidden rows/cols, error cells, mixed data types in columns, whitespace +
casing variants of category strings, cross-sheet formula references, cell
comments, color coding that carries semantic meaning, and so on.

Usage:
  python3 scripts/inspect-xlsx-deep.py path/to/file.xlsx [--sheets SHEET1,SHEET2]

Output: Markdown intended to be appended to the file's existing
`<filename> — MANIFEST.md` in `docs/REFLUJO/`, under a "## Deep findings"
heading. After running this, Jorge does a per-sheet visual inspection and
adds a "## Visual inspection" section. Only after both passes are in the
manifest does the parser get written.

Runtime deps: openpyxl 3.1+ (already installed for the structural manifest).
"""

from __future__ import annotations

import argparse
import sys
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ERROR_VALUES = {"#REF!", "#N/A", "#DIV/0!", "#NAME?", "#VALUE!", "#NULL!", "#NUM!", "#GETTING_DATA"}


def out(s: str = "") -> None:
    sys.stdout.write(s + "\n")


def fmt_cell_value(v: Any, max_len: int = 50) -> str:
    if v is None:
        return "∅"
    s = str(v)
    if len(s) > max_len:
        return repr(s[: max_len - 1] + "…")
    return repr(s)


# ─── Inspections ─────────────────────────────────────────────────────────────


def inspect_merged_cells(sheet: Worksheet) -> list[str]:
    """List every merged-cell range in the sheet."""
    findings = []
    for mr in sheet.merged_cells.ranges:
        # The top-left cell holds the value; the others are stripped.
        top_left = sheet.cell(row=mr.min_row, column=mr.min_col)
        findings.append(
            f"  - `{mr.coord}` ({mr.max_row - mr.min_row + 1}×{mr.max_col - mr.min_col + 1}) → {fmt_cell_value(top_left.value)}"
        )
    return findings


def inspect_hidden(sheet: Worksheet) -> tuple[list[int], list[str]]:
    """Hidden rows and hidden columns."""
    hidden_rows = [r for r, dim in sheet.row_dimensions.items() if dim.hidden]
    hidden_cols = [c for c, dim in sheet.column_dimensions.items() if dim.hidden]
    return hidden_rows, hidden_cols


def inspect_errors(sheet: Worksheet, max_report: int = 50) -> list[str]:
    """Cells holding formula errors (#REF!, #N/A, etc.) or whose formula evaluates to one."""
    findings = []
    for row in sheet.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and v.strip() in ERROR_VALUES:
                findings.append(f"  - `{cell.coordinate}`: {v.strip()}")
            elif cell.data_type == "e":
                findings.append(f"  - `{cell.coordinate}`: error cell ({v!r})")
            if len(findings) >= max_report:
                findings.append(f"  - … truncated at {max_report}")
                return findings
    return findings


def inspect_cross_sheet_formulas(sheet: Worksheet, this_sheet_name: str, max_report: int = 30) -> list[str]:
    """Formulas that reference other sheets. Captures dependencies the parser
    must resolve (or deliberately ignore)."""
    findings = []
    seen_refs: Counter[str] = Counter()
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type != "f":
                continue
            formula = str(cell.value or "")
            # Look for `'SheetName'!` or `SheetName!` patterns
            import re
            for m in re.finditer(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ ]*))!", formula):
                ref_name = m.group(1) or m.group(2)
                if ref_name != this_sheet_name:
                    seen_refs[ref_name] += 1
    for name, count in seen_refs.most_common():
        findings.append(f"  - References `{name}`: {count} formulas")
    if len(findings) > max_report:
        findings = findings[:max_report]
        findings.append(f"  - … truncated")
    return findings


def inspect_comments(sheet: Worksheet, max_report: int = 30) -> list[str]:
    """Cells with author notes / comments."""
    findings = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.comment is not None:
                text = cell.comment.text or ""
                author = cell.comment.author or "?"
                findings.append(f"  - `{cell.coordinate}` by {author}: {fmt_cell_value(text, max_len=80)}")
                if len(findings) >= max_report:
                    findings.append(f"  - … truncated at {max_report}")
                    return findings
    return findings


def inspect_column_type_mix(sheet: Worksheet, max_cols: int = 30, header_row: int = 1) -> list[str]:
    """For each column, count cell data types past the header. Catches
    "column is mostly numbers but has 3 string cells" — a parser-busting case."""
    findings = []
    max_col = min(sheet.max_column or 0, max_cols)
    for col_idx in range(1, max_col + 1):
        types: Counter[str] = Counter()
        sample_values_per_type: dict[str, list[str]] = {}
        for row_idx in range(header_row + 1, (sheet.max_row or 0) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            v = cell.value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            t = type(v).__name__
            types[t] += 1
            if t not in sample_values_per_type:
                sample_values_per_type[t] = []
            if len(sample_values_per_type[t]) < 2:
                sample_values_per_type[t].append(fmt_cell_value(v, max_len=30))
        if len(types) > 1:
            col_letter = get_column_letter(col_idx)
            mix = ", ".join(f"{t}×{n}" for t, n in types.most_common())
            samples = "; ".join(
                f"{t}: {', '.join(sample_values_per_type[t])}" for t in types
            )
            findings.append(f"  - **Col {col_letter}**: {mix}  ·  samples → {samples}")
    return findings


def inspect_whitespace_anomalies(sheet: Worksheet, max_report: int = 25) -> list[str]:
    """String cells with leading/trailing whitespace or double spaces.
    Common cause of "same value, multiple distinct strings" bugs."""
    findings = []
    for row in sheet.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v:
                continue
            if v != v.strip() or "  " in v or "\t" in v or "\n" in v:
                kind = []
                if v != v.lstrip():
                    kind.append("leading space")
                if v != v.rstrip():
                    kind.append("trailing space")
                if "  " in v:
                    kind.append("double space")
                if "\t" in v:
                    kind.append("tab")
                if "\n" in v:
                    kind.append("newline")
                findings.append(f"  - `{cell.coordinate}` [{', '.join(kind)}]: {fmt_cell_value(v, max_len=60)}")
                if len(findings) >= max_report:
                    findings.append(f"  - … truncated at {max_report}")
                    return findings
    return findings


def inspect_casing_variants(sheet: Worksheet, col_indices: list[int], max_per_col: int = 12) -> list[str]:
    """For each column likely to hold category strings, group unique values
    by their casefold+strip equivalent and flag any that have >1 surface form."""
    findings = []
    for col_idx in col_indices:
        groups: dict[str, set[str]] = {}
        for row_idx in range(2, (sheet.max_row or 0) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            v = cell.value
            if not isinstance(v, str) or not v.strip():
                continue
            key = v.strip().casefold()
            groups.setdefault(key, set()).add(v)
        variants = [(k, vs) for k, vs in groups.items() if len(vs) > 1]
        if variants:
            col_letter = get_column_letter(col_idx)
            findings.append(f"  - **Col {col_letter}** has {len(variants)} value(s) with multiple surface forms:")
            for _, vs in sorted(variants, key=lambda kv: -len(kv[1]))[:max_per_col]:
                findings.append(f"      - {' | '.join(sorted(repr(s) for s in vs))}")
    return findings


def inspect_color_coding(sheet: Worksheet, max_report: int = 30) -> list[str]:
    """Cells with non-default fill colors. May encode semantic meaning the
    parser needs to know about (e.g., red = void, yellow = needs review)."""
    findings = []
    color_counts: Counter[str] = Counter()
    sample_per_color: dict[str, str] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            fill = cell.fill
            if fill is None or fill.fill_type is None:
                continue
            try:
                rgb = fill.fgColor.rgb if fill.fgColor else None
            except AttributeError:
                rgb = None
            if not rgb or not isinstance(rgb, str):
                continue
            # Skip default white / no-fill
            if rgb.upper() in ("00000000", "FFFFFFFF", "FFFFFF", "00FFFFFF"):
                continue
            color_counts[rgb] += 1
            if rgb not in sample_per_color:
                sample_per_color[rgb] = f"{cell.coordinate}={fmt_cell_value(cell.value, max_len=30)}"
    for color, count in color_counts.most_common():
        findings.append(f"  - Fill `#{color}` × {count}  (e.g., {sample_per_color[color]})")
    return findings[:max_report]


def inspect_blank_row_islands(sheet: Worksheet, max_report: int = 20) -> list[str]:
    """Blank rows that interrupt a contiguous data range. Tells the parser
    whether to keep iterating past them or to treat them as boundaries."""
    findings = []
    in_data = False
    blank_streak_start: int | None = None
    seen_islands: list[tuple[int, int]] = []
    last_data_row = 0
    for row_idx in range(1, (sheet.max_row or 0) + 1):
        row_cells = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        is_blank = all(v is None or (isinstance(v, str) and not v.strip()) for v in row_cells)
        if not is_blank:
            if in_data and blank_streak_start is not None:
                seen_islands.append((blank_streak_start, row_idx - 1))
            in_data = True
            blank_streak_start = None
            last_data_row = row_idx
        elif in_data and blank_streak_start is None:
            blank_streak_start = row_idx
    for start, end in seen_islands[:max_report]:
        if start == end:
            findings.append(f"  - Row {start}: blank inside data range")
        else:
            findings.append(f"  - Rows {start}–{end}: blank streak inside data range ({end - start + 1} rows)")
    if len(seen_islands) > max_report:
        findings.append(f"  - … and {len(seen_islands) - max_report} more")
    return findings


def detect_pivot_tables(xlsx_path: Path) -> list[str]:
    """Pivot tables and Excel tables live in xl/pivotTables/ and xl/tables/
    inside the zipped xlsx. openpyxl's reader skips them; we inspect the zip."""
    findings = []
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            names = zf.namelist()
            pivots = sorted(n for n in names if n.startswith("xl/pivotTables/"))
            tables = sorted(n for n in names if n.startswith("xl/tables/"))
        if pivots:
            findings.append(f"  - **Pivot tables present:** {len(pivots)} → {', '.join(Path(p).name for p in pivots)}")
        if tables:
            findings.append(f"  - **Excel Tables (`<table>`):** {len(tables)} → {', '.join(Path(p).name for p in tables)}")
    except Exception as e:
        findings.append(f"  - _(error reading zip structure: {e})_")
    return findings


def inspect_freeze_pane_and_print_area(sheet: Worksheet) -> list[str]:
    findings = []
    if sheet.freeze_panes:
        findings.append(f"  - Freeze panes at `{sheet.freeze_panes}` (data likely starts at this row/col)")
    if sheet.print_area:
        findings.append(f"  - Print area: `{sheet.print_area}`")
    return findings


# ─── Driver ──────────────────────────────────────────────────────────────────


def inspect_sheet(sheet: Worksheet, sheet_name: str, xlsx_path: Path) -> None:
    out(f"\n## Deep findings — sheet: `{sheet_name}`\n")

    # Freeze panes + print area (cheap and informative)
    fp = inspect_freeze_pane_and_print_area(sheet)
    if fp:
        out("**Layout hints:**")
        for f in fp:
            out(f)
        out("")

    # Merged cells
    out("**Merged cells:**")
    mc = inspect_merged_cells(sheet)
    if mc:
        for f in mc[:50]:
            out(f)
        if len(mc) > 50:
            out(f"  - … and {len(mc) - 50} more merged ranges")
    else:
        out("  - None")
    out("")

    # Hidden rows / columns
    hidden_rows, hidden_cols = inspect_hidden(sheet)
    out("**Hidden rows / columns:**")
    out(f"  - Hidden rows: {hidden_rows if hidden_rows else 'none'}")
    out(f"  - Hidden columns: {hidden_cols if hidden_cols else 'none'}")
    out("")

    # Error cells
    out("**Error cells (`#REF!`, `#N/A`, etc.):**")
    err = inspect_errors(sheet)
    if err:
        for f in err:
            out(f)
    else:
        out("  - None")
    out("")

    # Cross-sheet formula refs
    out("**Cross-sheet formula references:**")
    csf = inspect_cross_sheet_formulas(sheet, sheet_name)
    if csf:
        for f in csf:
            out(f)
    else:
        out("  - None (or all self-references)")
    out("")

    # Cell comments
    out("**Cell comments / author notes:**")
    com = inspect_comments(sheet)
    if com:
        for f in com:
            out(f)
    else:
        out("  - None")
    out("")

    # Mixed-type columns
    out("**Columns with mixed data types** (past header row, type counts):")
    mtc = inspect_column_type_mix(sheet)
    if mtc:
        for f in mtc:
            out(f)
    else:
        out("  - Every column has a single consistent type")
    out("")

    # Whitespace anomalies
    out("**Whitespace / casing anomalies in string cells:**")
    ws = inspect_whitespace_anomalies(sheet)
    if ws:
        for f in ws:
            out(f)
    else:
        out("  - None")
    out("")

    # Blank-row islands
    out("**Blank rows inside data ranges:**")
    bri = inspect_blank_row_islands(sheet)
    if bri:
        for f in bri:
            out(f)
    else:
        out("  - None — data is contiguous")
    out("")

    # Color coding
    out("**Non-default cell fill colors** (may encode semantic meaning):")
    cc = inspect_color_coding(sheet)
    if cc:
        for f in cc:
            out(f)
    else:
        out("  - No notable color coding")
    out("")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("file", type=Path)
    ap.add_argument("--sheets", help="Comma-separated sheet names to inspect (default: all)")
    args = ap.parse_args()

    xlsx_path: Path = args.file
    if not xlsx_path.exists():
        sys.stderr.write(f"File not found: {xlsx_path}\n")
        return 2

    wb = load_workbook(xlsx_path, data_only=False)

    out(f"# Deep inspection — `{xlsx_path.name}`")
    out("")
    out("> Generated by `scripts/inspect-xlsx-deep.py`. This output is intended to be appended")
    out("> to the file's `<filename> — MANIFEST.md` under a top-level **`## Deep findings`**")
    out("> section. After it lands, Jorge does a per-sheet visual inspection and appends a")
    out("> **`## Visual inspection`** section. Only then is the parser written.")
    out("")

    # Workbook-level findings
    out("## Workbook-level")
    out("")
    pt = detect_pivot_tables(xlsx_path)
    if pt:
        out("**Pivot tables / Excel Tables (read via zip inspection):**")
        for f in pt:
            out(f)
    else:
        out("**Pivot tables / Excel Tables:** none detected")
    out("")
    out(f"**Sheets in workbook:** {len(wb.sheetnames)} → {', '.join(repr(s) for s in wb.sheetnames)}")
    out("")

    target_sheets = wb.sheetnames
    if args.sheets:
        wanted = [s.strip() for s in args.sheets.split(",")]
        target_sheets = [s for s in wb.sheetnames if s in wanted]
        if missing := [s for s in wanted if s not in wb.sheetnames]:
            sys.stderr.write(f"WARN: sheets not found: {missing}\n")

    for sheet_name in target_sheets:
        inspect_sheet(wb[sheet_name], sheet_name, xlsx_path)

    return 0


if __name__ == "__main__":
    sys.exit(main())
